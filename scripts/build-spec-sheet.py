"""Builds docs/Glimmora_Go_Spec.xlsx — single reference sheet listing every
role, page, feature, subfeature, and permission across the admin panel and
Kirana PWA. Derived from the SOW + Frontend Tracker + current codebase."""

import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
HDR_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="FEF3E7", end_color="FEF3E7", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="B45309")
STATUS_FILLS = {
    "Done":    PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "Partial": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "TODO":    PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "N/A":     PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
}
BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

def write_header(ws, cols, widths):
    for i, (col, width) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

def write_row(ws, row_idx, values, status_col=None, is_section=False):
    for i, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=i, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        if is_section:
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
    if status_col is not None and not is_section:
        status = values[status_col - 1]
        if status in STATUS_FILLS:
            ws.cell(row=row_idx, column=status_col).fill = STATUS_FILLS[status]

# ============================================================
# SHEET 1: Overview
# ============================================================
ws = wb.active
ws.title = "Overview"
ws["A1"] = "Glimmora Go — Spec Sheet"
ws["A1"].font = Font(bold=True, size=20, color="B45309")
ws["A2"] = "Frontend team: admin panel (Next.js) + Kirana PWA + public tracking"
ws["A2"].font = Font(size=11, color="475569")
ws["A3"] = "Generated from Glimmora_Go_SOW.docx + Frontend Tracker + current codebase"
ws["A3"].font = Font(size=10, italic=True, color="94A3B8")

ws["A5"] = "Status legend"
ws["A5"].font = Font(bold=True, size=11)
legend = [
    ("Done",    "Built, tested, in production-ready state"),
    ("Partial", "Skeleton exists but core sub-features missing"),
    ("TODO",    "Not yet built"),
    ("N/A",     "Owned by another team (backend/app-dev/infra) or deferred"),
]
for i, (status, desc) in enumerate(legend, start=6):
    c1 = ws.cell(row=i, column=1, value=status)
    c1.fill = STATUS_FILLS[status]
    c1.font = Font(bold=True)
    c1.border = BORDER
    c2 = ws.cell(row=i, column=2, value=desc)
    c2.alignment = WRAP
    c2.border = BORDER

ws["A12"] = "Sheets"
ws["A12"].font = Font(bold=True, size=11)
sheets_toc = [
    ("Roles",            "All 8 roles, access surface, primary responsibilities"),
    ("Permissions",      "Role × surface × read/write matrix"),
    ("Admin Panel",      "Every page with its features + subfeatures + status"),
    ("Kirana PWA",       "Every page with its features + subfeatures + status"),
    ("APIs",             "All HTTP endpoints with auth requirements"),
    ("Schema",           "Prisma models and what uses them"),
    ("Gaps",             "Items from the SOW not yet built"),
]
for i, (name, desc) in enumerate(sheets_toc, start=13):
    c1 = ws.cell(row=i, column=1, value=name)
    c1.font = Font(bold=True)
    c1.border = BORDER
    c2 = ws.cell(row=i, column=2, value=desc)
    c2.alignment = WRAP
    c2.border = BORDER

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 80

# ============================================================
# SHEET 2: Roles
# ============================================================
ws = wb.create_sheet("Roles")
write_header(
    ws,
    ["#", "Role", "Surface", "Seeded login", "Primary responsibilities", "Scoping", "Status"],
    [4, 18, 22, 28, 50, 24, 10],
)

roles = [
    (1, "Super Admin", "Admin Panel (full)", "admin@glimmora.ai / admin123",
     "Everything: cities, fares, coupons, drivers, rides, reports, admins, partners, referrals",
     "Global (all cities)", "Done"),
    (2, "Admin", "Admin Panel (full except admin user mgmt)", "admin2@glimmora.ai / admin123",
     "Same as Super Admin minus the ability to create/delete admin users",
     "Global (all cities)", "Done"),
    (3, "City Admin", "Admin Panel (scoped)", "rewa.admin@glimmora.ai / admin123",
     "Driver approvals, local fare, ride monitoring, reports, partner approvals — only for assigned city",
     "One cityId from Admin.cityId", "Done"),
    (4, "Verifier", "Admin Panel (drivers + docs only)", "verifier@glimmora.ai / admin123",
     "Approve/reject driver KYC documents. Sees only PENDING drivers. No access to other modules.",
     "All cities, PENDING drivers only", "Done"),
    (5, "Support / QA", "Admin Panel (read-only + complaints)", "support@glimmora.ai / admin123",
     "Monitor live rides, riders, reports, SOS alerts. (Ticketing module TODO.)",
     "Read-only global", "Done (no tickets module yet)"),
    (6, "Viewer", "Admin Panel (read-only)", "viewer@glimmora.ai / admin123",
     "Observer across all modules. Cannot write.",
     "Read-only global", "Done"),
    (7, "Kirana Agent", "Kirana PWA", "phone + 6-digit OTP",
     "Book rides on behalf of walk-in customers, view own bookings + commission, manage profile",
     "Own shop only (partnerId)", "Done"),
    (8, "Trusted Contact / Ride Recipient", "Public Tracking Link", "no login (opaque token URL)",
     "View one ride's live status + map. Used by riders' 'share-my-ride' and trusted-contact auto-SMS.",
     "One ride only via trackingToken", "Done"),
]
for i, r in enumerate(roles, start=2):
    write_row(ws, i, r, status_col=7)
ws.row_dimensions[1].height = 32

# ============================================================
# SHEET 3: Permissions (matrix)
# ============================================================
ws = wb.create_sheet("Permissions")
write_header(
    ws,
    ["Surface", "Super Admin", "Admin", "City Admin", "Verifier", "Support", "Viewer", "Notes"],
    [22, 12, 10, 14, 10, 10, 10, 40],
)

# W=write, R=read, — = no access
perm = [
    ("Dashboard",        "W", "W", "R (own city)", "—", "R", "R", "Stats + recent rides"),
    ("Rides",            "W", "W", "W (own city)", "—", "R", "R", "Cancel/complete/reassign/share"),
    ("SOS Alerts",       "R", "R", "R (own city)", "—", "R", "R", "Live sidebar badge"),
    ("Drivers list",     "W", "W", "W (own city)", "R (PENDING only)", "—", "R", ""),
    ("Driver docs",      "W", "W", "W (own city)", "W", "—", "—", "Per-document approve/reject + note"),
    ("Riders",           "R", "R", "R", "—", "R", "R", "Search + lifetime spend"),
    ("Fares",            "W", "W", "W (own city)", "—", "—", "R", "Per-city fare config"),
    ("Concessions",      "W", "W", "—", "—", "—", "R", "Women/senior/children multipliers"),
    ("Coupons",          "W", "W", "—", "—", "—", "R", "Flat/%, usage limit, expiry"),
    ("Cities",           "W", "W", "—", "—", "—", "R", "Includes archetype defaults editor"),
    ("Referrals",        "W", "W", "—", "—", "—", "R", "Auto-reward + manual override"),
    ("Reports",          "R", "R", "R (own city)", "—", "R", "R", "Includes CSV export + city breakdown"),
    ("Kirana Partners",  "W", "W", "W (own city)", "—", "—", "R", "Approve/reject + commission"),
    ("Admins",           "W", "—", "—", "—", "—", "—", "Super Admin only"),
    ("Track public",     "—", "—", "—", "—", "—", "—", "Public, opaque token"),
    ("Kirana PWA",       "—", "—", "—", "—", "—", "—", "Separate auth; agents only"),
]
for i, p in enumerate(perm, start=2):
    write_row(ws, i, p)

# ============================================================
# SHEET 4: Admin Panel (features + subfeatures)
# ============================================================
ws = wb.create_sheet("Admin Panel")
write_header(
    ws,
    ["#", "Page / Route", "Feature", "Sub-feature", "Access (roles)", "Status", "Notes"],
    [4, 28, 28, 46, 28, 10, 36],
)

admin_rows = [
    # Auth
    ("A", "AUTH", "", "", "", "", "", True),
    (1, "/login", "Admin login", "Email + password, JWT session cookie, redirect to ?next=", "public", "Done", ""),
    (2, "/api/auth/logout (POST)", "Logout", "Clear session cookie", "Authed", "Done", ""),

    # Dashboard
    ("A", "DASHBOARD", "", "", "", "", "", True),
    (3, "/", "Dashboard stats", "Rides today, revenue today, active drivers, SOS today", "SUPER/ADMIN/CITY/SUPP/VIEW", "Done", "CITY sees own-city only"),
    (4, "/", "Recent rides feed", "Top 8 most recent rides with links", "SUPER/ADMIN/CITY/SUPP/VIEW", "Done", ""),
    (5, "/", "DB-not-reachable banner", "Friendly message when Prisma can't connect", "all", "Done", ""),

    # Rides
    ("A", "RIDES", "", "", "", "", "", True),
    (6, "/rides", "Ride monitoring table", "100-row list, rider + driver + city + fare + status + started", "SUPER/ADMIN/CITY/SUPP/VIEW", "Done", ""),
    (7, "/rides", "Status filter pills", "ALL / REQUESTED / MATCHED / IN_TRIP / COMPLETED / CANCELLED", "all allowed", "Done", ""),
    (8, "/rides", "Live auto-refresh toggle", "15s polling, on/off, refresh counter", "all allowed", "Done", "Socket.io replacement-ready"),
    (9, "/rides", "Cancel ride", "PATCH action=CANCEL; blocked on terminal", "SUPER/ADMIN/CITY", "Done", ""),
    (10, "/rides", "Force-complete ride", "PATCH action=COMPLETE; sets completedAt, fareFinal", "SUPER/ADMIN/CITY", "Done", "Triggers referral reward"),
    (11, "/rides", "Reassign driver", "Inline select of APPROVED drivers in same city + PATCH", "SUPER/ADMIN/CITY", "Done", ""),
    (12, "/rides", "Share public link", "POST /api/rides/[id]/token → clipboard /track/[token]", "SUPER/ADMIN/CITY", "Done", ""),

    # SOS
    ("A", "SOS ALERTS", "", "", "", "", "", True),
    (13, "/sos", "SOS event log", "Filtered to sosTriggered=true rides, LIVE pulse on active", "SUPER/ADMIN/CITY/SUPP/VIEW", "Done", ""),
    (14, "/sos", "Stat cards", "Active SOS, SOS today, all-time count", "same", "Done", ""),
    (15, "Sidebar", "Live badge counter", "Polls /api/sos every 30s; red pill when > 0", "same", "Done", ""),

    # Drivers
    ("A", "DRIVERS", "", "", "", "", "", True),
    (16, "/drivers", "Driver list", "100-row list with status/online/docs/rides count", "SUPER/ADMIN/CITY/VERIF/VIEW", "Done", "VERIFIER sees PENDING only; CITY_ADMIN city-scoped"),
    (17, "/drivers", "Status filter pills", "ALL/PENDING/APPROVED/REJECTED/SUSPENDED", "same", "Done", ""),
    (18, "/drivers/[id]", "Driver profile card", "Status badge, city, phone, online, daily goal, rides, referral code", "SUPER/ADMIN/CITY/VERIF", "Done", ""),
    (19, "/drivers/[id]", "Approve / Reject / Suspend / Pending", "PATCH with optional verification note", "SUPER/ADMIN/CITY", "Done", ""),
    (20, "/drivers/[id]", "Verification note banner", "Shown when driver.verificationNote set", "same", "Done", ""),
    (21, "/drivers/[id]", "Document list + per-doc review", "Type, uploaded date, View file link, approve/reject + reviewNote", "SUPER/ADMIN/CITY/VERIF", "Done", "VERIFIER can only write docs"),
    ("—", "/drivers/[id]", "Document file upload UI", "Pick type + file, POST /api/uploads/driver-doc, stored under public/uploads/drivers/", "SUPER/ADMIN/CITY/VERIF", "Done", ""),

    # Riders
    ("A", "RIDERS", "", "", "", "", "", True),
    (22, "/riders", "Rider search + list", "Phone/name search, language badge, ride count, lifetime spend, last ride, joined", "SUPER/ADMIN/SUPP/VIEW", "Done", ""),
    ("—", "API", "Riders JSON API", "GET /api/riders (q/limit/offset), GET /api/riders/[id] with rides + stats", "riders:read", "Done", "For external clients and future rider app"),

    # Fares
    ("A", "FARES", "", "", "", "", "", True),
    (23, "/fares", "Per-city fare editor", "Base fare, per km, per min, minimum fare", "SUPER/ADMIN/CITY (write) · VIEW (read)", "Done", "CITY sees own city only"),
    (24, "/fares", "Saved ✓ feedback", "2s confirmation after PUT", "same", "Done", ""),

    # Concessions
    ("A", "CONCESSIONS", "", "", "", "", "", True),
    (25, "/concessions", "Per-city multipliers", "Women, senior, children — default 0.85/0.80/0.90", "SUPER/ADMIN (write) · VIEW (read)", "Done", ""),

    # Coupons
    ("A", "COUPONS", "", "", "", "", "", True),
    (26, "/coupons", "Create coupon", "Code, description, FLAT/PERCENT, amount, usage limit, expiry", "SUPER/ADMIN", "Done", ""),
    (27, "/coupons", "List + usage stats", "Used count / limit, expiry date, active/disabled/expired badge", "same + VIEW read", "Done", ""),
    (28, "/coupons", "Disable / Enable / Delete", "PATCH active flag / DELETE", "SUPER/ADMIN", "Done", ""),

    # Cities
    ("A", "CITIES", "", "", "", "", "", True),
    (29, "/cities", "Archetype defaults editor", "Per-archetype (Metro/Small Town) defaults: radius, surge, payment options, base, perKm, perMin, minFare", "SUPER/ADMIN (write) · VIEW (read)", "Done", "New cities inherit these"),
    (30, "/cities", "Add city form", "Name + state + archetype", "SUPER/ADMIN", "Done", ""),
    (31, "/cities", "Per-city edit row", "Archetype override, matching radius, surge, active toggle, payment pills", "SUPER/ADMIN", "Done", ""),

    # Referrals
    ("A", "REFERRALS", "", "", "", "", "", True),
    (32, "/referrals", "Stat cards", "Total / joined / rewarded with conversion %", "SUPER/ADMIN (write) · VIEW (read)", "Done", ""),
    (33, "/referrals", "Top referrers list", "Top 5 drivers by referredDrivers count", "same", "Done", ""),
    (34, "/referrals", "Recent referrals table", "Referee phone, rides done / 10, status (Pending/Joined/Rewarded), date", "same", "Done", ""),
    (35, "/referrals", "Issue / Revoke reward", "Manual toggle on Referral.rewardIssued", "SUPER/ADMIN", "Done", ""),
    (36, "/referrals", "Mark joined", "Manual toggle on refereeJoined", "SUPER/ADMIN", "Done", ""),
    (37, "/referrals", "Recompute rewards button", "Scans all drivers, grants rewards for any who crossed 10 rides", "SUPER/ADMIN", "Done", ""),
    (38, "automation", "Auto-reward on ride complete", "On PATCH action=COMPLETE, if driver was referred + hit 10 rides, extend referrer subscriptionUntil by 7 days", "system", "Done", ""),

    # Reports
    ("A", "REPORTS", "", "", "", "", "", True),
    (39, "/reports", "Range pills", "Last 7 / 14 / 30 / 90 days", "SUPER/ADMIN/CITY/SUPP/VIEW", "Done", ""),
    (40, "/reports", "Stat cards", "Rides, revenue, completion %, active drivers", "same", "Done", ""),
    (41, "/reports", "Charts", "Rides-per-day bar, revenue line, booking-channel donut", "same", "Done", "Chart.js"),
    (42, "/reports", "By-city breakdown", "Rides, completion %, revenue per city", "same", "Done", ""),
    (43, "/reports", "Top drivers table", "Top 10 by rides + revenue", "same", "Done", ""),
    (44, "/reports", "CSV export", "GET /api/reports/export?days=N → rides-Nd-YYYY-MM-DD.csv", "same", "Done", "City-filtered for CITY_ADMIN"),

    # Partners
    ("A", "KIRANA PARTNERS (admin side)", "", "", "", "", "", True),
    (45, "/partners", "Partner list with status filter", "Shop, owner, phone, city, commission %, bookings count, status + review note", "SUPER/ADMIN/CITY (write) · VIEW (read)", "Done", ""),
    (46, "/partners", "Approve / Reject / Suspend", "PATCH status + optional reviewNote", "SUPER/ADMIN/CITY", "Done", ""),
    (47, "/partners", "Commission % editor", "Per-partner override (default 10%)", "same", "Done", ""),

    # Subscriptions
    ("A", "SUBSCRIPTIONS", "", "", "", "", "", True),
    ("—", "/subscriptions", "Subscription list + stats", "All driver plans with active/expired/revoked status; active/total/revenue cards", "SUPER/ADMIN (write) · VIEW (read)", "Done", ""),
    ("—", "/subscriptions", "Grant subscription", "Pick approved driver + plan (DAILY/WEEKLY/MONTHLY) + amount; stacks on existing expiry", "SUPER/ADMIN", "Done", ""),
    ("—", "/subscriptions", "Revoke / Reinstate", "PATCH active flag; rolls back driver.subscriptionUntil if all plans inactive", "SUPER/ADMIN", "Done", ""),

    # Admins
    ("A", "ADMIN USER MGMT", "", "", "", "", "", True),
    (48, "/admins", "Create admin form", "Email, name, password (8+), role, city picker (if CITY_ADMIN)", "Super Admin only", "Done", ""),
    (49, "/admins", "Admin list", "Name, email, role badge, city, active/disabled, joined", "Super Admin only to see controls", "Done", "Non-super sees list read-only"),
    (50, "/admins", "Enable / Disable / Delete", "Self-delete + self-deactivate blocked", "Super Admin only", "Done", ""),

    # Cross-cutting
    ("A", "CROSS-CUTTING", "", "", "", "", "", True),
    (51, "sidebar", "Role-filtered nav", "Hide pages user can't access; role label under logo", "all", "Done", ""),
    (52, "layout", "Session enforcement", "requireAccess(surface) on every page; redirects unauthorized to /", "all", "Done", ""),
    (53, "API", "requireWrite + cityMismatch", "Centralized guards on every API mutation", "all", "Done", ""),
    (54, "global", "i18n Hindi + English toggle", "next-intl across all admin strings", "all", "TODO", "Tracker #20, Low priority"),
    (55, "global", "Audit log / compliance module", "Track who changed what, when", "SUPER/ADMIN", "TODO", "Post-MVP per SOW"),
    (56, "global", "Enterprise admin console", "Corporate/fleet accounts", "—", "N/A", "Post-MVP"),
    (57, "global", "Socket.io realtime rides", "Replace polling with live push", "—", "TODO", "Backend dependency"),
]

row_idx = 2
for item in admin_rows:
    if len(item) == 8 and item[-1] is True:
        # Section header
        merged = [""] + list(item[1:-1])
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c, value=item[1] if c == 1 else "")
        ws.cell(row=row_idx, column=1, value=item[1])
        for c in range(1, 8):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
            cell.border = BORDER
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    else:
        write_row(ws, row_idx, item, status_col=6)
    row_idx += 1

# ============================================================
# SHEET 5: Kirana PWA
# ============================================================
ws = wb.create_sheet("Kirana PWA")
write_header(
    ws,
    ["#", "Page / Route", "Feature", "Sub-feature", "Access", "Status", "Notes"],
    [4, 26, 28, 46, 26, 10, 36],
)

kirana_rows = [
    ("K", "PUBLIC (no session)", "", "", "", "", "", True),
    (1, "/k/signup", "Partner onboarding form", "Shop name, owner name, 10-digit phone, city picker", "public", "Done", ""),
    (2, "/k/signup", "Success screen", "'Application submitted' with link to /k/login", "public", "Done", ""),
    (3, "/k/login", "Phone → OTP request", "10-digit validation, 'Send OTP' triggers /api/kirana/otp/request", "public", "Done", ""),
    (4, "/k/login", "OTP entry screen", "6-digit code input, auto-surfaced in dev mode, back button", "public", "Done", ""),
    (5, "/k/login", "Verify + session", "/api/kirana/otp/verify sets glimmora_kirana_session cookie", "public", "Done", ""),

    ("K", "AUTHED (any status)", "", "", "", "", "", True),
    (6, "/k (any)", "Pending / Rejected / Suspended gate", "Blocks app with status-specific message + admin's review note + sign-out button", "Kirana partner (non-APPROVED)", "Done", ""),
    (7, "/k (logout)", "Sign out", "POST /api/kirana/logout clears cookie → redirect /k/login", "Kirana partner", "Done", ""),

    ("K", "APPROVED PARTNER", "", "", "", "", "", True),
    (8, "/k", "Dashboard today stats", "Today's bookings count, today's commission", "Approved partner", "Done", ""),
    (9, "/k", "Lifetime commission card", "Total earned on COMPLETED rides × commissionPct", "same", "Done", ""),
    (10, "/k", "CTA button", "'Book a ride for a customer' → /k/book", "same", "Done", ""),
    (11, "/k", "Recent 5 bookings", "Pickup → drop, status badge, date, fare", "same", "Done", ""),

    (12, "/k/book", "Customer section", "Rider phone (10-digit), optional name", "Approved partner", "Done", ""),
    (13, "/k/book", "Pickup section", "Address + lat + lng", "same", "Done", ""),
    ("—", "/k/book", "Map-based pickup/drop picker", "Leaflet + OSM Nominatim; click map or search address to set pickup and drop", "same", "Done", ""),
    (14, "/k/book", "Drop section", "Address + lat + lng", "same", "Done", ""),
    (15, "/k/book", "Concession selector", "None / Women / Senior / Children — applied to fare", "same", "Done", ""),
    (16, "/k/book", "Fare estimate", "PUT /api/kirana/bookings returns fare + distance + duration + your commission", "same", "Done", ""),
    (17, "/k/book", "Confirm → create ride", "POST /api/kirana/bookings creates Ride with bookingChannel=KIRANA, bookedByPartnerId=self", "same", "Done", ""),
    ("—", "/k/book", "SMS notification to rider", "Text rider their pickup code + ETA", "same", "TODO", "Needs MSG91"),
    ("—", "/k/book", "Auto-match driver", "Find nearest APPROVED online driver and assign", "same", "TODO", "No matching engine yet"),

    (18, "/k/bookings", "Bookings history list", "All partner's rides, status, rider phone, fare, commission earned", "Approved partner", "Done", ""),
    (19, "/k/bookings", "Commission per row", "Shown only when ride is COMPLETED", "same", "Done", ""),

    (20, "/k/profile", "Profile read-only", "Shop, owner, phone, city, commission %, joined, status badge", "Approved partner", "Done", ""),
    (21, "/k/profile", "Sign out button", "Same as dashboard logout", "same", "Done", ""),
    ("—", "/k/profile", "KYC document upload + status", "Pick type + file, upload to /api/kirana/documents; shows per-doc PENDING/APPROVED/REJECTED badge", "same", "Done", ""),

    ("K", "NAVIGATION + LAYOUT", "", "", "", "", "", True),
    (22, "layout", "Mobile-first shell", "Top header with shop name, bottom nav (Home/Book/History/Profile)", "Approved partner", "Done", ""),
    (23, "layout", "Session guard", "requireKirana() on every protected page", "all kirana routes", "Done", ""),

    ("K", "PWA SHELL", "", "", "", "", "", True),
    ("—", "manifest", "Web App Manifest", "public/manifest.webmanifest scoped to /k/ with name/icons/theme/start_url", "public", "Done", ""),
    ("—", "icons", "App icons + apple-touch", "192/512 PNG + maskable 512 + 180 apple-touch, brand orange 'GG' monogram", "public", "Done", "Generated via scripts/build-pwa-icons.py"),
    ("—", "service-worker", "Service worker (/sw.js, scope /k/)", "Cache-first static, network-first kirana APIs, offline fallback to /offline.html", "public", "Done", "Service-Worker-Allowed header set in next.config.js"),
    ("—", "install", "Install prompt component", "Captures beforeinstallprompt, shows 'Add to home screen' banner with dismiss persistence", "Approved partner", "Done", ""),
    ("—", "push", "Push notifications", "Driver matched / arrived / completed alerts", "system", "TODO", "SW ready; needs VAPID + backend trigger"),

    ("K", "CROSS-CUTTING", "", "", "", "", "", True),
    (24, "OTP", "6-digit, 5-min TTL", "OtpRequest table; single-use (usedAt timestamp)", "system", "Done", ""),
    (25, "OTP", "MSG91 send", "Actually send the code via SMS", "system", "TODO", "Stubbed; returns devCode in dev"),
    (26, "fare", "Haversine + concession", "src/lib/fare.ts shared with future rider/driver apps", "system", "Done", ""),
    (27, "commission", "Payout / settlement", "Actually paying partners their earned commission", "system", "TODO", "Tracks earned; no payout flow"),
    (29, "i18n", "Hindi + English toggle", "react-i18next across all kirana strings", "system", "TODO", "Low priority"),
]

row_idx = 2
for item in kirana_rows:
    if len(item) == 8 and item[-1] is True:
        for c in range(1, 8):
            cell = ws.cell(row=row_idx, column=c, value=item[1] if c == 1 else "")
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
            cell.border = BORDER
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    else:
        write_row(ws, row_idx, item, status_col=6)
    row_idx += 1

# ============================================================
# SHEET 6: APIs
# ============================================================
ws = wb.create_sheet("APIs")
write_header(
    ws,
    ["Method", "Path", "Purpose", "Auth / permission", "Status"],
    [10, 48, 46, 30, 10],
)
apis = [
    # Admin auth
    ("POST", "/api/auth/login",                 "Admin email+password login, sets session cookie", "public", "Done"),
    ("POST", "/api/auth/logout",                "Clear admin session",              "authed", "Done"),
    # Cities
    ("POST", "/api/cities",                     "Create city (uses archetype defaults)", "cities:write", "Done"),
    ("PATCH","/api/cities/[id]",                "Update archetype/radius/surge/active/payments", "cities:write", "Done"),
    ("PUT",  "/api/cities/[id]/fare",           "Update fare config",               "fares:write + city scope", "Done"),
    ("PUT",  "/api/cities/[id]/concessions",    "Update concession multipliers",    "concessions:write", "Done"),
    ("PUT",  "/api/archetypes/[archetype]",     "Edit Metro/Small Town defaults",   "cities:write", "Done"),
    # Drivers
    ("PATCH","/api/drivers/[id]",               "Approve/reject/suspend driver + note", "drivers:write + city scope", "Done"),
    ("GET",  "/api/drivers/[id]/documents/[docId]", "Fetch single driver document",     "documents:read + city scope", "Done"),
    ("PATCH","/api/drivers/[id]/documents/[docId]", "Approve/reject driver doc + review note", "documents:write + city scope", "Done"),
    ("DELETE","/api/drivers/[id]/documents/[docId]","Delete driver doc + file",         "documents:write + city scope", "Done"),
    ("POST", "/api/uploads/driver-doc",         "Upload driver doc file to storage",     "documents:write + city scope", "Done"),
    # Riders (read-only API)
    ("GET",  "/api/riders",                     "List riders (q/limit/offset)",         "riders:read", "Done"),
    ("GET",  "/api/riders/[id]",                "Rider detail with recent rides + stats","riders:read", "Done"),
    # Subscriptions
    ("GET",  "/api/subscriptions",              "List subscriptions (optional driverId/active filters)","subscriptions:read", "Done"),
    ("POST", "/api/subscriptions",              "Grant subscription (stacks on existing expiry)","subscriptions:write", "Done"),
    ("PATCH","/api/subscriptions/[id]",         "Revoke / reinstate subscription",      "subscriptions:write", "Done"),
    # Rides
    ("PATCH","/api/rides/[id]",                 "Cancel / Complete / Reassign driver", "rides:write + city scope", "Done"),
    ("POST", "/api/rides/[id]/token",           "Generate or fetch public tracking token", "rides:write + city scope", "Done"),
    # Coupons
    ("POST", "/api/coupons",                    "Create coupon",                    "coupons:write", "Done"),
    ("PATCH","/api/coupons/[id]",               "Toggle active",                    "coupons:write", "Done"),
    ("DELETE","/api/coupons/[id]",              "Delete coupon",                    "coupons:write", "Done"),
    # Referrals
    ("PATCH","/api/referrals/[id]",             "Manual reward issue/revoke",       "referrals:write", "Done"),
    ("POST", "/api/referrals/recompute",        "Recompute all pending referral rewards", "referrals:write", "Done"),
    # Reports
    ("GET",  "/api/reports/export",             "CSV download, city-filtered",      "reports:read", "Done"),
    # SOS
    ("GET",  "/api/sos",                        "Active SOS count (sidebar badge)", "sos:read + city filter", "Done"),
    # Admins
    ("POST", "/api/admins",                     "Create admin user",                "SUPER_ADMIN", "Done"),
    ("PATCH","/api/admins/[id]",                "Enable/disable, change role/city/password", "SUPER_ADMIN", "Done"),
    ("DELETE","/api/admins/[id]",               "Delete admin",                     "SUPER_ADMIN", "Done"),
    # Partners (admin side)
    ("PATCH","/api/partners/[id]",              "Approve/reject/suspend + commission %", "partners:write + city scope", "Done"),
    # Public track
    ("GET",  "/api/track/[token]",              "Public ride state for tracking page", "public", "Done"),
    # Kirana
    ("POST", "/api/kirana/signup",              "Partner application (creates PENDING)", "public", "Done"),
    ("POST", "/api/kirana/otp/request",         "Generate 6-digit OTP (phone must belong to a partner)", "public", "Done"),
    ("POST", "/api/kirana/otp/verify",          "Verify OTP + set kirana session cookie", "public", "Done"),
    ("POST", "/api/kirana/logout",              "Clear kirana session",             "kirana session", "Done"),
    ("PUT",  "/api/kirana/bookings",            "Fare estimate without creating",   "kirana session (any status)", "Done"),
    ("POST", "/api/kirana/bookings",            "Create booking (REQUESTED ride with bookedByPartnerId)", "kirana session (APPROVED only)", "Done"),
]
for i, a in enumerate(apis, start=2):
    write_row(ws, i, a, status_col=5)

# ============================================================
# SHEET 7: Schema
# ============================================================
ws = wb.create_sheet("Schema")
write_header(
    ws,
    ["Model / Enum", "Purpose", "Key fields", "Used by", "Status"],
    [24, 36, 46, 30, 10],
)
schema = [
    ("Admin",              "Admin panel users",           "email, name, passwordHash, role, cityId", "Login, /admins page", "Done"),
    ("AdminRole (enum)",   "6 roles",                     "SUPER_ADMIN, ADMIN, CITY_ADMIN, VERIFIER, SUPPORT, VIEWER", "RBAC matrix", "Done"),
    ("Rider",              "Ride customers",              "phone (unique), name, language", "/riders, kirana signup auto-creates", "Done"),
    ("Driver",              "Fleet drivers",              "phone, name, cityId, status, online, lat/lng, referredById, referralRewardGranted", "/drivers", "Done"),
    ("DriverStatus (enum)","Driver KYC states",           "PENDING, APPROVED, REJECTED, SUSPENDED", "driver detail", "Done"),
    ("DriverDocument",     "KYC docs per driver",         "driverId, type, fileUrl, status, reviewNote", "driver detail + upload UI", "Done"),
    ("DocumentType (enum)","Doc categories",              "LICENSE, RC, INSURANCE, AADHAAR, PAN", "driver docs", "Done"),
    ("DocumentStatus (enum)","Doc review states",         "PENDING, APPROVED, REJECTED", "driver docs", "Done"),
    ("City",               "Cities served",               "name, state, archetype, matchingRadiusKm, surgeMultiplier, paymentOptions, active", "/cities, fare/concession, partner/driver scoping", "Done"),
    ("CityArchetype (enum)","Tier",                       "METRO, SMALL_TOWN", "Drives defaults + UI", "Done"),
    ("ArchetypeDefaults",  "Tier defaults",               "archetype (PK), radius, surge, payments, baseFare, perKm, perMin, minFare", "/cities archetype editor, applied on new-city create", "Done"),
    ("FareConfig",         "Per-city fare rules",         "cityId (unique), baseFare, perKm, perMin, minimumFare, women/senior/children multipliers", "/fares, /concessions", "Done"),
    ("Ride",               "Trips",                       "riderId, driverId, cityId, pickup/drop, fareEstimate, fareFinal, status, bookingChannel, trackingToken, bookedByPartnerId, sosTriggered", "/rides, /track, /sos, /reports, kirana bookings", "Done"),
    ("RideStatus (enum)",  "Lifecycle",                   "REQUESTED, MATCHED, EN_ROUTE, ARRIVED, IN_TRIP, COMPLETED, CANCELLED", "Everywhere", "Done"),
    ("BookingChannel",     "Where a ride came from",      "APP, SMS, WHATSAPP, KIRANA, IVR", "Reports, kirana rides", "Done"),
    ("ConcessionType",     "Fare discount category",      "NONE, WOMEN, SENIOR, CHILDREN", "Fare calc", "Done"),
    ("Coupon",             "Promo codes",                 "code (unique), discountType, amount, usageLimit, usedCount, validUntil, active", "/coupons", "Done"),
    ("DiscountType",       "Coupon type",                 "FLAT, PERCENT", "coupons", "Done"),
    ("Referral",           "Driver-to-driver referrals",  "referrerDriverId, referrerRiderId, refereePhone, refereeJoined, ridesCompleted, rewardIssued", "/referrals manual controls", "Done"),
    ("Subscription",       "Driver sachet packs",         "driverId, plan, amount, startedAt, expiresAt, active", "/subscriptions admin + referral reward extends this", "Done (admin grant; no driver-side purchase UI)"),
    ("SubscriptionPlan",   "Sachet durations",            "DAILY, WEEKLY, MONTHLY", "", "Done"),
    ("KiranaPartner",      "Agent/kirana partners",       "phone (unique), shopName, ownerName, cityId, commissionPct, status, reviewNote", "/partners, Kirana PWA", "Done"),
    ("PartnerStatus",      "Partner KYC states",          "PENDING, APPROVED, REJECTED, SUSPENDED", "/partners", "Done"),
    ("PartnerDocument",    "Partner KYC docs",            "partnerId, type, fileUrl, status", "Kirana /k/profile upload + /partners/[id] review", "Done"),
    ("OtpRequest",         "Phone OTP codes",             "phone, code, purpose, expiresAt, usedAt", "Kirana OTP flow", "Done"),
    ("OtpPurpose",         "OTP use cases",               "KIRANA_LOGIN, RIDER_LOGIN, DRIVER_LOGIN", "Kirana now; rider/driver later", "Done (only Kirana used)"),
]
for i, s in enumerate(schema, start=2):
    write_row(ws, i, s, status_col=5)

# ============================================================
# SHEET 8: Gaps
# ============================================================
ws = wb.create_sheet("Gaps")
write_header(
    ws,
    ["#", "Gap", "Module", "Owner per SOW", "Blocks", "Priority", "Notes"],
    [4, 40, 22, 22, 36, 12, 30],
)
gaps = [
    (1,  "Rider React Native / PWA app",               "Rider",         "App Dev team", "Every rider-side feature", "P0", "SOW says RN; pragmatic swap to PWA"),
    (2,  "Driver React Native / PWA app",              "Driver",        "App Dev team", "Driver accept/location/earnings/docs", "P0", ""),
    (3,  "Matching engine (nearest online driver)",    "Backend",       "Backend team", "Auto-assignment on ride create", "P0", "src/lib/fare.ts has Haversine; add /api/rides/match"),
    (4,  "MSG91 SMS gateway",                          "Integrations",  "Backend team", "OTP SMS, fare estimate, job push, receipts, safety", "P0", ""),
    (5,  "WhatsApp Business API (Gupshup/Wati)",       "Integrations",  "Backend team", "WhatsApp lite booking bot, status, receipts", "P1", ""),
    (6,  "Scheduled ride matcher",                     "Backend",       "Backend team", "Pre-booked rides auto-match at T-15min", "P1", "Vercel Cron can do this"),
    (7,  "Offline driver SQLite + SMS job accept",     "Backend+RN",    "Backend + App Dev", "Patchy-2G driver workflow", "P1", "Big effort; RN only"),
    (8,  "Panic phrase voice detection",               "RN",            "App Dev", "Silent SOS", "P2", "RN Voice + on-device keyword listener"),
    (9,  "Trusted-contact SMS on ride start",          "Backend",       "Backend team", "Auto-alert to emergency contact", "P1", ""),
    (10, "Socket.io live tracking server",             "Backend",       "Backend team", "Real-time position updates in /rides + /track", "P2", "Currently polling"),
    (11, "Payment (Razorpay)",                         "Integrations",  "Backend team", "UPI/card fares, driver subscriptions", "P1", "SOW says cash-only for MVP; fast-follow"),
    (12, "Driver subscription purchase flow",          "Driver PWA",    "App Dev + Backend", "Actually charging drivers", "P1", "Schema exists"),
    (13, "Driver referral share UI",                   "Driver PWA",    "App Dev", "Drivers referring drivers in-app", "P1", "Backend reward logic is done"),
    (14, "Driver daily earnings goal screen",          "Driver PWA",    "App Dev", "Progress bar + celebration", "P2", ""),
    (15, "Driver demand heat map",                     "Driver PWA",    "App Dev", "Show clustering of rides to drivers", "P2", ""),
    (16, "Rider share-my-ride button",                 "Rider PWA",     "App Dev", "User-facing trigger for /track/:token", "P1", "Token API + page already exist"),
    (17, "Push notifications",                         "RN/Backend",    "App Dev + Backend", "Ride updates, promos", "P2", ""),
    (18, "KYC document upload UI (drivers + partners)","Admin + PWA",   "Frontend + Backend", "Actually getting documents uploaded", "Done", "Shipped — stored under public/uploads/ in dev; production should swap to Vercel Blob / S3"),
    (19, "i18n Hindi + English",                       "All frontends", "Frontend + App Dev", "Vernacular UX", "P3", "Low priority per tracker"),
    (20, "Complaints / tickets module",                "Admin",         "Frontend", "Support role has nothing to operate on", "P2", "Support role already scoped in RBAC"),
    (21, "Audit log",                                  "Admin",         "Frontend + Backend", "Compliance, who-changed-what", "P2", "Post-MVP per SOW"),
    (22, "Enterprise / corporate admin console",       "Admin",         "Frontend", "B2B customer fleets", "P3", "Post-MVP"),
    (23, "Commission payout flow",                     "Admin + PWA",   "Backend + Frontend", "Actually paying partners", "P2", "Earnings tracked; no payout"),
    (24, "Map-based pickup/drop picker (Kirana)",      "Kirana PWA",    "Frontend", "Stop asking agents for lat/lng", "Done", "Shipped — Leaflet + Nominatim"),
    (25, "IVR / missed-call booking",                  "Integrations",  "Backend", "Non-smartphone rider channel", "P3", "SOW explicitly deferred to Phase 2"),
    (26, "Driver subscription purchase UI",             "Driver PWA",    "App Dev + Backend", "Drivers buying their own plan; admin grant exists", "P1", "Admin /subscriptions shipped; waits on payments + Driver app"),
    (27, "VAPID push notifications",                    "RN/PWA + Backend","App Dev + Backend", "Booking status push to kirana partners", "P2", "PWA service worker ready; needs backend trigger + VAPID keys"),
    (28, "Bookings detail + live track in Kirana PWA",  "Kirana PWA",    "Frontend", "Partner can't watch a ride they booked", "P1", "Token API + public /track page already exist"),
    (29, "Per-document DELETE in partner docs review",  "Admin",         "Frontend", "Admins can approve/reject but not remove partner docs", "P3", ""),
]
for i, g in enumerate(gaps, start=2):
    write_row(ws, i, g)

# Save
os.makedirs("docs", exist_ok=True)
out = "docs/Glimmora_Go_Spec.xlsx"
wb.save(out)
print(f"Wrote {out}")
